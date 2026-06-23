import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def generate_excel_report(df: pd.DataFrame, output_path: str) -> None:
    """
    Gera um relatório executivo final em formato .xlsx com identidade visual
    corporativa, formatação financeira e fórmulas nativas do Excel.
    """
    print(f"\nIniciando a geração do relatório estilizado em: {output_path}")
    
    # Criar diretório de saída caso não exista
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    
    # 1. Instanciar Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Relatório Executivo"
    
    # Garantir que as linhas de grade (gridlines) fiquem visíveis no Excel
    ws.views.sheetView[0].showGridLines = True
    
    # 2. Paleta de Cores e Estilos (Identidade Visual Grafite/Azul Corporativo)
    NAVY_BLUE = "1F497D"      # Cor principal para cabeçalhos e títulos
    LIGHT_BLUE = "DCE6F1"     # Cor para a linha de totais
    LIGHT_GRAY = "F2F5F8"     # Cor para linhas zebradas
    CARD_BG = "F2F5F9"        # Cor de fundo dos cards de KPI
    BORDER_COLOR = "D9D9D9"   # Cor cinza claro para as bordas internas
    
    # Fontes
    font_title = Font(name="Segoe UI", size=16, bold=True, color=NAVY_BLUE)
    font_subtitle = Font(name="Segoe UI", size=10, italic=True, color="595959")
    font_kpi_label = Font(name="Segoe UI", size=9, bold=True, color="595959")
    font_kpi_val = Font(name="Segoe UI", size=14, bold=True, color=NAVY_BLUE)
    font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    font_data = Font(name="Segoe UI", size=11)
    font_total = Font(name="Segoe UI", size=11, bold=True, color="000000")
    
    # Fills
    fill_header = PatternFill(start_color=NAVY_BLUE, end_color=NAVY_BLUE, fill_type="solid")
    fill_zebra = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")
    fill_total = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
    fill_kpi = PatternFill(start_color=CARD_BG, end_color=CARD_BG, fill_type="solid")
    
    # Alignments
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    
    # Borders
    thin_side = Side(style="thin", color=BORDER_COLOR)
    thick_side = Side(style="medium", color=NAVY_BLUE)
    double_bottom = Side(style="double", color="000000")
    thin_top = Side(style="thin", color="000000")
    
    border_cell = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    border_total = Border(left=thin_side, right=thin_side, top=thin_top, bottom=double_bottom)
    border_kpi = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    
    # 3. Título e Cabeçalho do Relatório
    ws.merge_cells("A2:I2")
    ws["A2"] = "RELATÓRIO EXECUTIVO DE VENDAS"
    ws["A2"].font = font_title
    ws["A2"].alignment = align_center
    ws.row_dimensions[2].height = 25
    
    ws.merge_cells("A3:I3")
    ws["A3"] = "Consolidação e análise automática de múltiplos canais de venda"
    ws["A3"].font = font_subtitle
    ws["A3"].alignment = align_center
    ws.row_dimensions[3].height = 18
    
    # 4. Determinar intervalos da tabela principal
    start_row = 10
    n_rows = len(df)
    end_row = start_row + n_rows - 1
    total_row = end_row + 1
    
    # 5. Cards de KPI no Topo (Linhas 5 e 6)
    # Card 1: Faturamento Total
    ws.merge_cells("B5:C5")
    ws.merge_cells("B6:C6")
    ws["B5"] = "FATURAMENTO TOTAL"
    ws["B6"] = f"=SUM(G{start_row}:G{end_row})"  # Injeta fórmula nativa
    
    # Card 2: Ticket Médio por Transação
    ws.merge_cells("E5:F5")
    ws.merge_cells("E6:F6")
    ws["E5"] = "TICKET MÉDIO (POR VENDA)"
    ws["E6"] = f"=AVERAGE(G{start_row}:G{end_row})"  # Injeta fórmula nativa
    
    # Card 3: Volume total de transações
    ws.merge_cells("H5:I5")
    ws.merge_cells("H6:I6")
    ws["H5"] = "VOLUME DE VENDAS"
    ws["H6"] = f"=COUNTA(A{start_row}:A{end_row})"  # Injeta fórmula nativa
    
    # Estilizando os Cards de KPI
    kpi_ranges = [("B5", "B6", "C5", "C6"), ("E5", "E6", "F5", "F6"), ("H5", "H6", "I5", "I6")]
    for label_cell_ref, val_cell_ref, end_col_l, end_col_v in kpi_ranges:
        ws[label_cell_ref].font = font_kpi_label
        ws[label_cell_ref].alignment = align_center
        ws[val_cell_ref].font = font_kpi_val
        ws[val_cell_ref].alignment = align_center
        
        # Aplicar formato de número nos KPIs relevantes
        if label_cell_ref == "B5" or label_cell_ref == "E5":
            ws[val_cell_ref].number_format = '"R$ " #,##0.00'
        else:
            ws[val_cell_ref].number_format = '#,##0'
            
        # Aplicar borda e fundo nas células mescladas dos cards
        cols_to_border = [label_cell_ref[0], end_col_l]
        for col_name in cols_to_border:
            ws[f"{col_name}5"].fill = fill_kpi
            ws[f"{col_name}5"].border = border_kpi
            ws[f"{col_name}6"].fill = fill_kpi
            ws[f"{col_name}6"].border = border_kpi
            
    ws.row_dimensions[5].height = 15
    ws.row_dimensions[6].height = 22
    
    # 6. Cabeçalhos da Tabela (Linha 9)
    headers = [
        "ID Venda", "Data", "Canal de Venda", "Produto", 
        "Quantidade", "Valor Unitário", "Valor Total", 
        "ID Cliente", "Arquivo Origem"
    ]
    
    ws.row_dimensions[9].height = 24
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=9, column=col_idx)
        cell.value = header
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = border_cell
        
    # 7. Dados da Tabela (Linha 10 em diante)
    for row_idx, (_, row_data) in enumerate(df.iterrows(), start_row):
        ws.row_dimensions[row_idx].height = 20
        is_even = (row_idx % 2 == 0)
        row_fill = fill_zebra if is_even else PatternFill(fill_type=None)
        
        # Inserir valores e aplicar estilos célula a célula
        # Coluna A: ID Venda
        cell_id = ws.cell(row=row_idx, column=1, value=row_data['id_venda'])
        cell_id.alignment = align_center
        cell_id.number_format = '0'
        
        # Coluna B: Data
        cell_date = ws.cell(row=row_idx, column=2, value=row_data['data'].date())
        cell_date.alignment = align_center
        cell_date.number_format = 'yyyy-mm-dd'
        
        # Coluna C: Canal de Venda
        cell_channel = ws.cell(row=row_idx, column=3, value=row_data['canal'])
        cell_channel.alignment = align_left
        
        # Coluna D: Produto
        cell_product = ws.cell(row=row_idx, column=4, value=row_data['produto'])
        cell_product.alignment = align_left
        
        # Coluna E: Quantidade
        cell_qty = ws.cell(row=row_idx, column=5, value=row_data['quantidade'])
        cell_qty.alignment = align_right
        cell_qty.number_format = '#,##0'
        
        # Coluna F: Valor Unitário
        cell_unit = ws.cell(row=row_idx, column=6, value=row_data['valor_unitario'])
        cell_unit.alignment = align_right
        cell_unit.number_format = '"R$ " #,##0.00'
        
        # Coluna G: Valor Total (Fórmula Excel: Quantidade * Valor Unitário)
        cell_total = ws.cell(row=row_idx, column=7, value=f"=E{row_idx}*F{row_idx}")
        cell_total.alignment = align_right
        cell_total.number_format = '"R$ " #,##0.00'
        
        # Coluna H: ID Cliente
        cell_cust = ws.cell(row=row_idx, column=8, value=row_data['id_cliente'])
        cell_cust.alignment = align_center
        
        # Coluna I: Arquivo Origem
        cell_file = ws.cell(row=row_idx, column=9, value=row_data['arquivo_origem'])
        cell_file.alignment = align_left
        cell_file.font = Font(name="Segoe UI", size=9, color="7F7F7F") # Fonte menor e cinza
        
        # Aplicar fundo e borda para todas as células da linha
        for col_idx in range(1, 10):
            c = ws.cell(row=row_idx, column=col_idx)
            if col_idx != 9:  # Não sobrescrever fonte personalizada da origem
                c.font = font_data
            if row_fill.fill_type:
                c.fill = row_fill
            c.border = border_cell

    # 8. Linha de Totais da Tabela (total_row)
    ws.row_dimensions[total_row].height = 22
    
    # Mesclar colunas A a D para escrever "Total Geral"
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=4)
    ws.cell(row=total_row, column=1, value="Total Geral").font = font_total
    ws.cell(row=total_row, column=1).alignment = align_right
    
    # Injetar fórmulas de soma
    # Soma de Quantidades (Coluna E)
    cell_total_qty = ws.cell(row=total_row, column=5, value=f"=SUM(E{start_row}:E{end_row})")
    cell_total_qty.font = font_total
    cell_total_qty.alignment = align_right
    cell_total_qty.number_format = '#,##0'
    
    # Soma de Faturamento Total (Coluna G)
    cell_total_rev = ws.cell(row=total_row, column=7, value=f"=SUM(G{start_row}:G{end_row})")
    cell_total_rev.font = font_total
    cell_total_rev.alignment = align_right
    cell_total_rev.number_format = '"R$ " #,##0.00'
    
    # Estilizar a linha de totais completa
    for col_idx in range(1, 10):
        c = ws.cell(row=total_row, column=col_idx)
        c.fill = fill_total
        c.border = border_total
        
    # 9. Ajustar automaticamente a largura das colunas para evitar truncation/###
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        
        for cell in col:
            if cell.row < 9:
                # Ignorar as primeiras linhas (títulos e KPIs) no cálculo de largura
                continue
                
            if cell.value is not None:
                val_str = str(cell.value)
                # Se for uma fórmula do Excel, estimamos o tamanho
                if val_str.startswith('='):
                    if col_letter == 'G': # Valor Total com R$
                        max_len = max(max_len, 16)
                    elif col_letter == 'E': # Quantidade
                        max_len = max(max_len, 8)
                else:
                    # Adiciona margem para formatação financeira (R$ 1.000,00 tem mais caracteres que 1000)
                    if col_letter in ['F', 'G'] and isinstance(cell.value, (int, float)):
                        max_len = max(max_len, len(f"R$ {cell.value:,.2f}"))
                    else:
                        max_len = max(max_len, len(val_str))
                        
        # Aplica a largura calculada com uma margem de segurança
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
        
    # Salvar arquivo
    wb.save(output_path)
    print(f"Relatório gerado e salvo com sucesso em: {output_path}")
